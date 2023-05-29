from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from aerosandbox import XFoil, Airfoil
import numpy as np
import pandas as pd
# from pyfiglet import Figlet
from typing import List
from tqdm import tqdm


class QuickFoil:
    def __init__(self, airfoils: List[str], Re: float = 100000, Mach: float = 0, alphas: np.ndarray = np.arange(-5, 15, 0.25), sort_by: str = "CL") -> None:
        self.airfoils = airfoils
        if Re <= 0:
            raise ValueError("Reynolds number must be greater than 0")
        if (Mach > 0.3) or (Mach < 0):
            raise ValueError("Mach number must be 0 <= Mach <= 0.3")

        self.Re = Re
        self.Mach = Mach
        self.alphas = alphas
        self.sort_by = sort_by

    def run_xfoil(self) -> List:
        data = []
        progress = tqdm(self.airfoils)
        for airfoil in progress:
            progress.set_postfix_str(airfoil) # displays a progress bar with name of current airfoil
            # print(f"Now running {airfoil}...")
            xf_airfoil = Airfoil(name=airfoil)
            xf = XFoil(
                airfoil=xf_airfoil.repanel(n_points_per_side=200),
                Re = self.Re,
                mach=self.Mach,
                max_iter=1000,
                timeout=60,)
            result = xf.alpha(self.alphas)
            data.append(result)
            # print(f"Finished {airfoil}")
        self.data = data
        return data

    def write_excel(self, file_name='SimpFoil_Run') -> None:
        """
        CHANGE FILE PATH HERE. LEAVE \{file_name}.xlsx AND CHANGE EVERYTHING ELSE
        """
        
        file_path = fr'./{file_name}.xlsx'
        # workbook = Workbook()
        writer = pd.ExcelWriter(path=file_path, engine='openpyxl')
        dataFrame_list = []
        if self.sort_by in ["CL", "CL/CD"]:
            reverse_sort = True
        elif self.sort_by in ['CD', 'CDp', 'CM', 'Cpmin']:
            reverse_sort = False

        for i in range(len(self.data[0]['alpha'])):
            writing_data = []
            for j, airfoil in enumerate(self.data):
                writing_data.append({
                    "Airfoil": self.airfoils[j],
                    "CL": airfoil['CL'][i],
                    "CD": airfoil['CD'][i],
                    "CDp": airfoil['CDp'][i],
                    "CL/CD": airfoil['CL'][i]/airfoil['CD'][i],
                    "CM": airfoil['CM'][i],
                    "Cpmin": airfoil['Cpmin'][i],
                    "Top_Xtr": airfoil['Top_Xtr'][i],
                    "Bot_Xtr": airfoil['Bot_Xtr'][i]
                })
            
            dataFrame_list.append(pd.DataFrame(sorted(writing_data, key=lambda s: s[self.sort_by],reverse=reverse_sort)))

        
        alphas = [int(n) for n in np.ndarray.tolist(self.data[0]['alpha'])]

        for i, df in enumerate(dataFrame_list):
            df.to_excel(writer, index=False, header=True, sheet_name=f"Alpha(deg)={alphas[i]}")
        writer.close()
# def get_airfoils():
#     airfoils = []
#     while True:
#         try:
#             request = input("Airfoil: ")
#             if request.lower() == 'done':
#                 if len(airfoils) < 2:
#                     raise ValueError
#                 return airfoils
#             airfoils.append((request, Airfoil(name=request).repanel(150)))
#         except TypeError:
#             print("Invalid Airfoil! (Not a NACA 4 Digit or in UIUC Database)")
#         except ValueError:
#             print("Please input at least two airfoils")

# def get_Re():
#     while True:
#         try:
#             Re = float(input("Re: "))
#             return Re
#         except ValueError:
#             print("Invalid Reynolds Number")
#             continue

# def get_Mach():
#     while True:
#         try:
#             Mach = float(input("Mach: "))
#             if 0 < Mach <= 0.3:
#                 return Mach
#             else:
#                 raise ValueError    
#         except ValueError:
#             print("Invalid Mach Number")
#             continue

# def get_alphas():
#     while True:
#         try:
#             alpha_first = int(input("First Alpha:  "))
#             alpha_last = int(input("Last Alpha: "))
#             alpha_increment = float(input("Alpha Increment: "))
#             if alpha_last < alpha_first or alpha_increment < 0:
#                 raise ValueError

#             alphas = np.arange(alpha_first, alpha_last + 1, alpha_increment)            
#             return alphas
#         except ValueError:
#             print("Invalid Angle of Attack Range")
#             continue

# def get_optimizer():
#     while True:
#         optimizer = input("Optimize for: ").upper()
#         if optimizer in ['CL', 'CD', 'CDp', 'CL/CD', 'CM', 'Cpmin']:
#             return optimizer
        
# def get_fileName():
#     FileName = input("File Name: ")
#     if '.' not in FileName:
#         return FileName

# def run_XFoil(airfoil, re, mach, alphas):  
#         xf = XFoil(
#         airfoil=airfoil.repanel(n_points_per_side=200),
#         Re = re,
#         mach=mach,
#         max_iter=1000,
#         timeout=60,
#         )

        # return xf.alpha(alphas)

# def write_excel(airfoils, data, file_name='SimpFoil_Run',optimizer='CL'):
#     """
#     CHANGE FILE PATH HERE. LEAVE \{file_name}.xlsx AND CHANGE EVERYTHING ELSE
#     """
#     file_path = fr'C:\Users\reach\Desktop\2D_Airfoil_Easy_Anaysis_Project\Quick-and-Easy-2D-Airfoil-Analysis\Excel_Data\{file_name}.xlsx'
#     workbook = Workbook()
#     writer = pd.ExcelWriter(path=file_path, engine='openpyxl')
    
#     dataFrame_list = []

#     if optimizer in ["CL", "CL/CD"]:
#         reverse_sort = True
#     elif optimizer in ['CD', 'CDp', 'CM', 'Cpmin']:
#         reverse_sort = False

#     for i in range(len(data[0]['alpha'])):
#         writing_data = []
#         for j, airfoil in enumerate(data):
#             writing_data.append({
#                 "Airfoil": airfoils[j][0],
#                 "CL": airfoil['CL'][i],
#                 "CD": airfoil['CD'][i],
#                 "CDp": airfoil['CDp'][i],
#                 "CL/CD": airfoil['CL'][i]/airfoil['CD'][i],
#                 "CM": airfoil['CM'][i],
#                 "Cpmin": airfoil['Cpmin'][i],
#                 "Top_Xtr": airfoil['Top_Xtr'][i],
#                 "Bot_Xtr": airfoil['Bot_Xtr'][i]
#             })
        
#         dataFrame_list.append(pd.DataFrame(sorted(writing_data, key=lambda s: s[optimizer],reverse=reverse_sort)))

    
#     alphas = [int(n) for n in np.ndarray.tolist(data[0]['alpha'])]

#     for i, df in enumerate(dataFrame_list):
#         df.to_excel(writer, index=False, header=True, sheet_name=f"Alpha(deg)={alphas[i]}")
        

#     writer.close()


def main():
    airfoils = [
        "naca6409",
        "naca4412",]
        # "s4320",
        # "s4110",]
        # "sa7038",
        # "s7075",
        # "naca4415",
        # "naca2410",
        # "naca2410",
        # "naca2408",
        # "naca4424",
        # "naca2414",
        # "naca2415",
        # "naca1408",
        # "naca1410",
        # "naca1412",]
    
    Re = 500000.0
    Mach = 0.0
    alphas = np.arange(-5, 15, 0.5)
    sort_by = "CL"
    qf = QuickFoil(airfoils, Re=Re, Mach=Mach, alphas=alphas, sort_by=sort_by)

    data = qf.run_xfoil()
    # print(data)
    qf.write_excel(file_name="test")
    # figlet = Figlet()
    # figlet.setFont(font='doom')
    # print(figlet.renderText("Welcome to QuickFoil !"))

    # print("Enter the desired airfoils and type 'done' when finished")
    # airfoils = get_airfoils()
    # print()

    # print("Enter the desired Reynolds number")
    # Re = get_Re()
    # print()

    # print("Enter the desired Mach number")
    # Mach = get_Mach()
    # print()

    # print("Enter the desired range of angle of attacks")
    # alphas = get_alphas()
    # print()

    # Optimizer = get_optimizer()
    # print()

    # print("Enter the desired file name")
    # FileName = get_fileName()

    # data = []
    # for airfoil in airfoils :  
    #     data.append(run_XFoil(airfoil[1], Re, Mach, alphas))

    # write_excel(airfoils, data, file_name=FileName, optimizer=Optimizer)

    # print(figlet.renderText("Success !"))

if __name__ == '__main__':
    main()